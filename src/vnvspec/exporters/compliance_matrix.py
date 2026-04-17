"""Compliance matrix exporter — maps requirements to standard clauses.

Produces XLSX, CSV, or HTML showing clause-level coverage with evidence verdicts.
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Literal

from vnvspec.core.assessment import Report
from vnvspec.core.trace import standard_gap_analysis


def export_compliance_matrix(
    report: Report,
    *,
    standard: str,
    spec: Any,
    path: Path | str,
    output_format: Literal["xlsx", "csv", "html"] = "xlsx",
) -> Path:
    """Export a compliance matrix to the given format.

    Parameters
    ----------
    report:
        The assessment report containing evidence.
    standard:
        Registry name (e.g. ``"iso_pas_8800"``).
    spec:
        The :class:`~vnvspec.core.spec.Spec` being assessed.
    path:
        Output file path.
    output_format:
        One of ``"xlsx"``, ``"csv"``, ``"html"``.

    Returns
    -------
    Path
        The written file path.
    """
    path = Path(path)
    gap_report = standard_gap_analysis(spec, standard)

    # Build evidence lookup: requirement_id -> list of (evidence_id, verdict)
    evidence_map: dict[str, list[tuple[str, str]]] = {}
    for ev in report.evidence:
        evidence_map.setdefault(ev.requirement_id, []).append((ev.id, ev.verdict))

    # Build rows
    rows: list[dict[str, str]] = []
    for clause in gap_report.clauses:
        ev_ids: list[str] = []
        verdicts: list[str] = []
        for req_id in clause.mapped_requirements:
            for ev_id, verdict in evidence_map.get(req_id, []):
                ev_ids.append(ev_id)
                verdicts.append(verdict)

        rows.append(
            {
                "Clause ID": clause.clause_id,
                "Clause": clause.clause,
                "Title": clause.title,
                "Normative Level": clause.normative_level,
                "Mapped Requirements": ", ".join(clause.mapped_requirements),
                "Evidence IDs": ", ".join(ev_ids),
                "Verdicts": ", ".join(verdicts),
                "Coverage Status": clause.status,
            }
        )

    if output_format == "xlsx":
        _write_xlsx(rows, path, gap_report.standard)
    elif output_format == "csv":
        _write_csv(rows, path)
    elif output_format == "html":
        _write_html(rows, path, gap_report.standard)

    return path


def _write_xlsx(rows: list[dict[str, str]], path: Path, title: str) -> None:
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

    wb = Workbook()
    ws = wb.active
    if ws is None:  # pragma: no cover
        ws = wb.create_sheet()
    ws.title = "Compliance Matrix"

    # Header
    headers = list(rows[0].keys()) if rows else []
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    status_colors = {
        "covered": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "partial": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        "gap": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    for row_idx, row in enumerate(rows, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[header])
            if header == "Coverage Status":
                fill = status_colors.get(row[header])
                if fill:
                    cell.fill = fill

    # Auto-fit column widths (approximate)
    from openpyxl.utils import get_column_letter  # noqa: PLC0415

    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in rows:
            max_len = max(max_len, len(row[header]))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(path)


def _write_csv(rows: list[dict[str, str]], path: Path) -> None:
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def _write_html(rows: list[dict[str, str]], path: Path, title: str) -> None:
    headers = list(rows[0].keys()) if rows else []
    status_css = {
        "covered": "background-color: #C6EFCE;",
        "partial": "background-color: #FFEB9C;",
        "gap": "background-color: #FFC7CE;",
    }

    html_parts = [
        "<!DOCTYPE html>",
        f"<html><head><title>Compliance Matrix — {title}</title>",
        "<style>table{border-collapse:collapse;width:100%}th,td{border:1px solid #ccc;"
        "padding:8px;text-align:left}th{background:#4472C4;color:white}</style>",
        "</head><body>",
        f"<h1>Compliance Matrix — {title}</h1>",
        "<table><thead><tr>",
    ]
    for h in headers:
        html_parts.append(f"<th>{h}</th>")
    html_parts.append("</tr></thead><tbody>")

    for row in rows:
        style = status_css.get(row.get("Coverage Status", ""), "")
        html_parts.append("<tr>")
        for h in headers:
            cell_style = f' style="{style}"' if h == "Coverage Status" and style else ""
            html_parts.append(f"<td{cell_style}>{row[h]}</td>")
        html_parts.append("</tr>")

    html_parts.append("</tbody></table></body></html>")
    path.write_text("\n".join(html_parts), encoding="utf-8")
