"""Tests for the compliance matrix exporter."""

from __future__ import annotations

from pathlib import Path

import pytest

from vnvspec.core.assessment import Report
from vnvspec.core.evidence import Evidence
from vnvspec.core.requirement import Requirement
from vnvspec.core.spec import Spec
from vnvspec.exporters.compliance_matrix import export_compliance_matrix


@pytest.fixture()
def sample_spec_with_standards() -> Spec:
    return Spec(
        name="compliance-test",
        requirements=[
            Requirement(
                id="REQ-001",
                statement="The system shall define safety goals.",
                rationale="Compliance.",
                verification_method="analysis",
                acceptance_criteria=["Goals documented."],
                standards={"iso_pas_8800": ["6.2.1"]},
            ),
            Requirement(
                id="REQ-002",
                statement="The system shall identify hazards.",
                rationale="Safety.",
                verification_method="analysis",
                acceptance_criteria=["Hazards listed."],
                standards={"iso_pas_8800": ["6.2.2"]},
            ),
        ],
    )


@pytest.fixture()
def sample_report() -> Report:
    return Report(
        spec_name="compliance-test",
        evidence=[
            Evidence(id="EV-001", requirement_id="REQ-001", kind="analysis", verdict="pass"),
            Evidence(id="EV-002", requirement_id="REQ-002", kind="analysis", verdict="fail"),
        ],
    )


class TestComplianceMatrixXLSX:
    def test_xlsx_output(
        self,
        sample_spec_with_standards: Spec,
        sample_report: Report,
        tmp_path: Path,
    ) -> None:
        out = tmp_path / "matrix.xlsx"
        result = export_compliance_matrix(
            sample_report,
            standard="iso_pas_8800",
            spec=sample_spec_with_standards,
            path=out,
            output_format="xlsx",
        )
        assert result == out
        assert out.exists()
        assert out.stat().st_size > 0

    def test_xlsx_cell_content(
        self,
        sample_spec_with_standards: Spec,
        sample_report: Report,
        tmp_path: Path,
    ) -> None:
        from openpyxl import load_workbook

        out = tmp_path / "matrix.xlsx"
        export_compliance_matrix(
            sample_report,
            standard="iso_pas_8800",
            spec=sample_spec_with_standards,
            path=out,
        )
        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None
        # Check header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
        assert "Clause ID" in headers
        assert "Coverage Status" in headers


class TestComplianceMatrixCSV:
    def test_csv_output(
        self,
        sample_spec_with_standards: Spec,
        sample_report: Report,
        tmp_path: Path,
    ) -> None:
        out = tmp_path / "matrix.csv"
        result = export_compliance_matrix(
            sample_report,
            standard="iso_pas_8800",
            spec=sample_spec_with_standards,
            path=out,
            output_format="csv",
        )
        assert result == out
        assert out.exists()
        content = out.read_text()
        assert "Clause ID" in content
        assert "Coverage Status" in content


class TestComplianceMatrixHTML:
    def test_html_output(
        self,
        sample_spec_with_standards: Spec,
        sample_report: Report,
        tmp_path: Path,
    ) -> None:
        out = tmp_path / "matrix.html"
        result = export_compliance_matrix(
            sample_report,
            standard="iso_pas_8800",
            spec=sample_spec_with_standards,
            path=out,
            output_format="html",
        )
        assert result == out
        assert out.exists()
        content = out.read_text()
        assert "<table>" in content
        assert "Compliance Matrix" in content
